"""Выгрузка готовой сводной в оформленный xlsx.

Файл уходит поставщику, поэтому оформление здесь — не украшательство, а часть
задачи. Разъезжающиеся колонки, «19867.45» рядом с «3160.1000000000004» и
отсутствие шапки в прежнем отчёте появлялись не по небрежности: сводная Excel
отдаёт лист без форматов, а править его руками раз в месяц никто не станет.

Всё, что задавалось руками, задаётся здесь один раз: двухэтажная шапка с
объединением по акции, числовые форматы, ширины по содержимому, закрепление
области, автофильтр, выделенная строка итога, подписи и настройки печати.
"""
from __future__ import annotations

import os
from datetime import date, datetime
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import Cell, Metric, ReportTable

FONT = "Calibri"

# Палитра сдержанная: отчёт читают в печати и в чужой почте, где яркая заливка
# выглядит хуже, чем отсутствие всякой.
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
GROUP_FILL = PatternFill("solid", fgColor="2E5496")
TOTAL_FILL = PatternFill("solid", fgColor="DCE3F0")
STRIPE_FILL = PatternFill("solid", fgColor="F4F6FB")

HAIR = Side(style="hair", color="B4BFD4")
THIN = Side(style="thin", color="8EA0C0")
GRID = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
HEAD_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

QTY_FORMAT = "#,##0"
MONEY_FORMAT = "#,##0.00"

# Ширина колонки в знаках. Потолок нужен номенклатуре: названия доходят до
# полутора сотен символов, и по содержимому колонка уехала бы за край листа.
MIN_WIDTH = 11
MAX_WIDTH = 52
TEXT_PADDING = 3


def save(table: ReportTable, destination: str, *, author: str = "",
         today: date | None = None) -> str:
    """Записывает отчёт. Возвращает путь к готовому файлу."""
    if not destination:
        raise ValueError("Не указан файл для сохранения")
    book = Workbook()
    sheet = book.active
    sheet.title = table.period.title or "Отчёт"
    _sheet(sheet, table, author=author, today=today or date.today())
    if table.profile.stores_sheet and table.store_totals:
        _stores_sheet(book.create_sheet("По магазинам"), table)
    if directory := os.path.dirname(destination):
        os.makedirs(directory, exist_ok=True)
    book.save(destination)
    return destination


def default_name(table: ReportTable, folder: str) -> str:
    """Путь по шаблону имени из профиля: «…/Июнь 2026.xlsx»."""
    return os.path.join(folder, f"{_safe(table.file_name())}.xlsx")


def _safe(name: str) -> str:
    """Имя файла без символов, запрещённых в Windows."""
    cleaned = "".join(" " if char in '\\/:*?"<>|' else char for char in name)
    return " ".join(cleaned.split()) or "Отчёт"


# --- основной лист --------------------------------------------------------------

def _sheet(sheet: Worksheet, table: ReportTable, *, author: str, today: date) -> None:
    width = len(table.row_fields) + table.column_count
    row = _header_block(sheet, table, width=width, author=author, today=today)
    head_top = row + 1
    row = _table_head(sheet, table, head_top)
    first_data = row
    row = _body(sheet, table, row)
    _total_row(sheet, table, row)
    _finish(sheet, table, head_top=head_top, first_data=first_data,
            last_row=row, width=width)


def _header_block(sheet: Worksheet, table: ReportTable, *, width: int,
                  author: str, today: date) -> int:
    """Шапка над таблицей: что за отчёт, за какой период и по каким магазинам."""
    lines: list[tuple[str, bool]] = [(table.header_title(), True)]
    if period := table.period.range():
        first, last = period
        lines.append((f"Период: {first:%d.%m.%Y} — {last:%d.%m.%Y}", False))
    if table.stores:
        lines.append((f"Магазинов: {len(table.stores)} · {_join(table.stores)}", False))
    if table.profile.note:
        lines.append((table.profile.note, False))
    stamp = f"Сформирован {today:%d.%m.%Y}"
    lines.append((f"{stamp} · {author}" if author else stamp, False))

    for index, (text, is_title) in enumerate(lines, start=1):
        cell = sheet.cell(row=index, column=1, value=text)
        cell.font = Font(name=FONT, size=14 if is_title else 9, bold=is_title,
                         color="1F3864" if is_title else "5A6784")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if width > 1:
            sheet.merge_cells(start_row=index, start_column=1,
                              end_row=index, end_column=width)
        sheet.row_dimensions[index].height = 22 if is_title else 14
    return len(lines) + 1


def _table_head(sheet: Worksheet, table: ReportTable, top: int) -> int:
    """Двухэтажная шапка: акция сверху, метрики под ней.

    Когда разрез колонок не задан, второй этаж не нужен — шапка становится
    однострочной, и лишняя пустая строка в файле не появляется.
    """
    grouped = bool(table.column_fields)
    bottom = top + 1 if grouped else top

    for index, role in enumerate(table.row_fields, start=1):
        cell = sheet.cell(row=top, column=index, value=role.title)
        _style_head(cell, GROUP_FILL)
        if grouped:
            sheet.merge_cells(start_row=top, start_column=index,
                              end_row=bottom, end_column=index)
            _style_head(sheet.cell(row=bottom, column=index), GROUP_FILL)

    column = len(table.row_fields) + 1
    span = len(table.metrics)
    for group in table.groups:
        if grouped:
            cell = sheet.cell(row=top, column=column, value=group.label)
            _style_head(cell, GROUP_FILL)
            if span > 1:
                sheet.merge_cells(start_row=top, start_column=column,
                                  end_row=top, end_column=column + span - 1)
            for offset in range(span):
                _style_head(sheet.cell(row=top, column=column + offset), GROUP_FILL)
        for offset, metric in enumerate(table.metrics):
            cell = sheet.cell(row=bottom, column=column + offset, value=metric.title)
            _style_head(cell, HEADER_FILL)
        column += span

    sheet.row_dimensions[top].height = 30
    if grouped:
        sheet.row_dimensions[bottom].height = 30
    return bottom + 1


def _style_head(cell, fill: PatternFill) -> None:
    cell.fill = fill
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = HEAD_BORDER


def _body(sheet: Worksheet, table: ReportTable, top: int) -> int:
    row = top
    for position, item in enumerate(table.rows):
        stripe = position % 2 == 1
        for index, key in enumerate(item.keys, start=1):
            cell = sheet.cell(row=row, column=index, value=key)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=True)
            cell.border = GRID
            if stripe:
                cell.fill = STRIPE_FILL
        for offset, data in enumerate(item.cells):
            _value(sheet.cell(row=row, column=len(item.keys) + 1 + offset), data,
                   stripe=stripe)
        row += 1
    return row


def _value(cell, data: Cell, *, stripe: bool = False, bold: bool = False) -> None:
    """Число в ячейку.

    Ноль записывается пустотой намеренно: в разрезе по акциям большинство
    пересечений пустые, и сетка из нулей читается хуже, чем пробел. Значение
    округляется при записи — иначе в файл попадает «3160.1000000000004», как
    это было в прежнем отчёте.
    """
    if not data.empty:
        cell.value = round(data.value, 2 if data.metric.money else 3)
    cell.number_format = MONEY_FORMAT if data.metric.money else QTY_FORMAT
    cell.font = Font(name=FONT, size=10, bold=bold)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = GRID
    if bold:
        cell.fill = TOTAL_FILL
    elif stripe:
        cell.fill = STRIPE_FILL


def _total_row(sheet: Worksheet, table: ReportTable, row: int) -> None:
    label = sheet.cell(row=row, column=1, value="Итого")
    label.font = Font(name=FONT, size=10, bold=True)
    label.alignment = Alignment(horizontal="left", vertical="center")
    label.fill = TOTAL_FILL
    label.border = GRID
    for index in range(2, len(table.row_fields) + 1):
        filler = sheet.cell(row=row, column=index)
        filler.fill = TOTAL_FILL
        filler.border = GRID
    for offset, data in enumerate(table.totals):
        _value(sheet.cell(row=row, column=len(table.row_fields) + 1 + offset),
               data, bold=True)


def _finish(sheet: Worksheet, table: ReportTable, *, head_top: int,
            first_data: int, last_row: int, width: int) -> None:
    """Ширины, закрепление, фильтр, подписи и печать."""
    _widths(sheet, table, head_top=head_top, first_data=first_data,
            last_row=last_row)

    # Закрепление по нижней строке шапки и по колонкам-разрезам: в отчёте на
    # три акции прокрутка вправо иначе уводит название товара за край.
    sheet.freeze_panes = sheet.cell(row=first_data, column=len(table.row_fields) + 1)
    head_bottom = first_data - 1
    # Фильтр ставится по данным, но без строки итога: попав в диапазон, она
    # уезжает вместе с отфильтрованными строками и перестаёт быть итогом.
    if last_row > first_data:
        sheet.auto_filter.ref = (
            f"A{head_bottom}:{get_column_letter(width)}{last_row - 1}")

    _signatures(sheet, table, last_row + 2, width)

    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    # Шапка повторяется на каждой печатной странице: отчёт на две сотни позиций
    # без этого читается только с первой страницы.
    sheet.print_title_rows = f"{head_top}:{head_bottom}"


def _widths(sheet: Worksheet, table: ReportTable, *, head_top: int,
            first_data: int, last_row: int) -> None:
    """Ширина по самому длинному значению колонки, с потолком и полом.

    Заголовки шапки в расчёт не берутся целиком: «Скидка бонусами, руб» шире
    любого числа под ним, и колонка растянулась бы впустую — она переносится
    по словам.
    """
    for index in range(1, len(table.row_fields) + table.column_count + 1):
        longest = 0
        for row in range(first_data, last_row + 1):
            value = sheet.cell(row=row, column=index).value
            if value is None:
                continue
            text = (f"{value:,.2f}" if isinstance(value, float)
                    else str(value))
            longest = max(longest, len(text))
        if index <= len(table.row_fields):
            longest = max(longest, len(table.row_fields[index - 1].title))
        else:
            metric = table.metrics[(index - len(table.row_fields) - 1)
                                   % len(table.metrics)]
            longest = max(longest, _head_width(metric))
        width = min(MAX_WIDTH, max(MIN_WIDTH, longest + TEXT_PADDING))
        sheet.column_dimensions[get_column_letter(index)].width = width


def _head_width(metric: Metric) -> int:
    """Шапка переносится по словам, поэтому колонке нужно самое длинное слово."""
    return max((len(word) for word in metric.title.split()), default=MIN_WIDTH)


def _signatures(sheet: Worksheet, table: ReportTable, row: int, width: int) -> None:
    """Подписи под таблицей. Пустой список — не повод рисовать пустой блок."""
    lines = [line for line in table.profile.signatures if line.strip()]
    if not lines:
        return
    for offset, text in enumerate(lines):
        cell = sheet.cell(row=row + offset, column=1, value=text)
        cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if width > 1:
            sheet.merge_cells(start_row=row + offset, start_column=1,
                              end_row=row + offset, end_column=width)


# --- лист «По магазинам» --------------------------------------------------------

def _stores_sheet(sheet: Worksheet, table: ReportTable) -> None:
    """Итоги по магазинам после применения правил объединения.

    Лист нужен менеджерам, которые собирают отчёт из точечных выгрузок: он
    показывает, сколько в итоге учтено за каждой точкой, и с ним расхождение с
    исходником находится за минуту, а не за вечер.
    """
    title = sheet.cell(row=1, column=1, value="Итоги по магазинам")
    title.font = Font(name=FONT, size=12, bold=True, color="1F3864")
    sheet.merge_cells(start_row=1, start_column=1,
                      end_row=1, end_column=len(table.metrics) + 1)

    head = 3
    _style_head(sheet.cell(row=head, column=1, value="Магазин"), GROUP_FILL)
    for offset, metric in enumerate(table.metrics):
        _style_head(sheet.cell(row=head, column=2 + offset, value=metric.title),
                    HEADER_FILL)
    sheet.row_dimensions[head].height = 30

    row = head + 1
    for position, (store, cells) in enumerate(table.store_totals):
        stripe = position % 2 == 1
        name = sheet.cell(row=row, column=1, value=store)
        name.font = Font(name=FONT, size=10)
        name.alignment = Alignment(horizontal="left", vertical="center")
        name.border = GRID
        if stripe:
            name.fill = STRIPE_FILL
        for offset, data in enumerate(cells):
            _value(sheet.cell(row=row, column=2 + offset), data, stripe=stripe)
        row += 1

    total = sheet.cell(row=row, column=1, value="Итого")
    total.font = Font(name=FONT, size=10, bold=True)
    total.fill = TOTAL_FILL
    total.border = GRID
    for offset, metric in enumerate(table.metrics):
        summed = sum(cells[offset].value for _, cells in table.store_totals)
        _value(sheet.cell(row=row, column=2 + offset), Cell(metric, summed), bold=True)

    sheet.column_dimensions["A"].width = min(MAX_WIDTH, max(
        MIN_WIDTH, max((len(name) for name, _ in table.store_totals), default=0) + TEXT_PADDING))
    for offset, metric in enumerate(table.metrics):
        sheet.column_dimensions[get_column_letter(2 + offset)].width = max(
            MIN_WIDTH, _head_width(metric) + TEXT_PADDING)
    sheet.freeze_panes = sheet.cell(row=head + 1, column=2)
    sheet.sheet_view.showGridLines = False


def _join(names: Sequence[str], limit: int = 4) -> str:
    """Список магазинов для шапки: длинный перечень сворачивается.

    Строка шапки объединена по всей ширине таблицы и не переносится, поэтому
    перечислять два десятка точек здесь нельзя — хвост просто обрежется.
    """
    if len(names) <= limit:
        return ", ".join(names)
    return ", ".join(names[:limit]) + f" и ещё {len(names) - limit}"
