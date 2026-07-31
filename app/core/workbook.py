"""Чтение и запись Excel-файлов с кэшированием по времени изменения."""
from __future__ import annotations

import io
import os
from typing import Any, Callable, Iterable, Sequence

import openpyxl

from .models import Column, FieldRole, Record, Sheet
from .normalize import (
    code_key,
    comparable,
    detect_noise_tokens,
    digits_only,
    extract_quantity,
    normalize_text,
    parse_quantity,
)
from .schema import clean_value, detect_columns, detect_header_row

ProgressCallback = Callable[[int, int], None]

# Роли, для которых нормализованное представление считается заранее и переиспользуется.
_NORMALIZED_ROLES = (
    FieldRole.NAME, FieldRole.NAME_ALT, FieldRole.BRAND, FieldRole.CATEGORY,
    FieldRole.COLOR, FieldRole.SIZE, FieldRole.MANUFACTURER,
    FieldRole.DESCRIPTION, FieldRole.NOTE, FieldRole.VOLUME,
)
_CODE_ROLES = (FieldRole.ARTICLE, FieldRole.SKU)

_cache: dict[tuple[str, str, float, int], Sheet] = {}

# Книга Excel нового формата — это ZIP-архив.
_ZIP_MAGIC = b"PK\x03\x04"


def open_workbook(
    path: str,
    *,
    data_only: bool = True,
    read_only: bool = False,
) -> openpyxl.Workbook:
    """Открывает книгу, не полагаясь на расширение файла.

    1С выгружает шаблон переоценки с расширением `.xls`, а внутри это обычный
    xlsx. openpyxl отказывается открывать такой файл по имени — проверка идёт
    по расширению, — поэтому содержимое передаётся ему потоком. Настоящий
    старый формат отсекается по сигнатуре: его openpyxl не читает вовсе, и
    пользователю нужно понятное объяснение, а не трассировка.
    """
    name = os.path.basename(path)
    try:
        with open(path, "rb") as handle:
            head = handle.read(len(_ZIP_MAGIC))
    except OSError as error:
        raise ValueError(f"Не удалось открыть файл {name}: {error}") from error
    if head != _ZIP_MAGIC:
        raise ValueError(
            f"{name}: файл в старом формате Excel. Откройте его в Excel "
            "и сохраните как «Книга Excel (*.xlsx)».")
    if path.lower().endswith((".xlsx", ".xlsm")):
        return openpyxl.load_workbook(path, read_only=read_only, data_only=data_only)
    with open(path, "rb") as handle:
        stream = io.BytesIO(handle.read())
    return openpyxl.load_workbook(stream, read_only=read_only, data_only=data_only)


def list_sheets(path: str) -> list[str]:
    workbook = open_workbook(path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_raw(
    path: str, sheet_name: str | None = None, limit: int | None = None
) -> tuple[list[list[Any]], str]:
    """Сырые строки листа и его имя — для предпросмотра и определения структуры."""
    workbook = open_workbook(path, read_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.worksheets[0]
        rows: list[list[Any]] = []
        for index, row in enumerate(worksheet.iter_rows(values_only=True)):
            if limit is not None and index >= limit:
                break
            rows.append([clean_value(v) for v in row])
        return rows, worksheet.title
    finally:
        workbook.close()


def load_sheet(
    path: str,
    sheet_name: str | None = None,
    header_row: int | None = None,
    role_overrides: dict[int, FieldRole] | None = None,
    progress: ProgressCallback | None = None,
) -> Sheet:
    """Загружает лист целиком: колонки, записи и предрассчитанные представления."""
    key = _cache_key(path, sheet_name, header_row, role_overrides)
    if (cached := _cache.get(key)) is not None:
        if progress:
            progress(len(cached.records), len(cached.records))
        return cached

    rows, resolved_name = read_raw(path, sheet_name)
    if not rows:
        raise ValueError(f"Лист не содержит данных: {os.path.basename(path)}")

    header_index = detect_header_row(rows) if header_row is None else header_row
    header = rows[header_index] if header_index < len(rows) else []
    body = [row for row in rows[header_index + 1:] if any(v is not None for v in row)]

    columns = detect_columns(header, body)
    for index, role in (role_overrides or {}).items():
        if 0 <= index < len(columns):
            columns[index].role = role
            columns[index].auto = False

    name_columns = [c.index for c in columns if c.role in (FieldRole.NAME, FieldRole.NAME_ALT)]
    noise = detect_noise_tokens(
        (" ".join(str(row[i]) for i in name_columns if i < len(row) and row[i] is not None) for row in body)
    )

    records = _build_records(body, columns, noise, header_index, progress)
    sheet = Sheet(
        path=path,
        sheet_name=resolved_name,
        header_row=header_index,
        columns=columns,
        records=records,
        noise_tokens=noise,
    )
    _cache[key] = sheet
    return sheet


def _build_records(
    body: Sequence[Sequence[Any]],
    columns: Sequence[Column],
    noise: frozenset[str],
    header_index: int,
    progress: ProgressCallback | None,
) -> list[Record]:
    roles = [(c.index, c.role) for c in columns if c.role is not FieldRole.OTHER]
    total = len(body)
    records: list[Record] = []
    for offset, row in enumerate(body):
        values = list(row)
        by_role = {
            role: values[index]
            for index, role in roles
            if index < len(values) and values[index] is not None
        }
        record = Record(row=header_index + offset + 2, values=values, by_role=by_role)
        prepare_record(record, noise)
        records.append(record)
        if progress and (offset % 250 == 0 or offset == total - 1):
            progress(offset + 1, total)
    return records


def prepare_record(record: Record, noise: frozenset[str]) -> None:
    """Считает нормализованные ключи один раз, чтобы поиск не пересчитывал их.

    Вызывается и из импортёров переоценки: они строят записи сами, зная
    структуру файла, но представления для сравнения должны считаться так же.
    """
    normalized = record.normalized
    for role in _NORMALIZED_ROLES:
        if (value := record.by_role.get(role)) is not None:
            normalized[role] = normalize_text(value)
    for role in _CODE_ROLES:
        if (value := record.by_role.get(role)) is not None:
            normalized[role] = code_key(value)
    if (ean := record.by_role.get(FieldRole.EAN)) is not None:
        normalized[FieldRole.EAN] = digits_only(ean)

    name = record.by_role.get(FieldRole.NAME)
    record.quantity = parse_quantity(record.by_role.get(FieldRole.VOLUME)) or extract_quantity(name)
    record.match_key = comparable(name, noise) if name is not None else ""
    record.search_blob = " ".join(dict.fromkeys(v for v in normalized.values() if v))


def _cache_key(
    path: str,
    sheet_name: str | None,
    header_row: int | None,
    role_overrides: dict[int, FieldRole] | None,
) -> tuple[str, str, float, int]:
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = 0.0
    overrides = tuple(sorted((k, v.value) for k, v in (role_overrides or {}).items()))
    return path, sheet_name or "", mtime, hash((header_row, overrides))


def invalidate_cache(path: str | None = None) -> None:
    if path is None:
        _cache.clear()
        return
    for key in [k for k in _cache if k[0] == path]:
        del _cache[key]


def write_sheet(
    destination: str,
    sheet_name: str,
    titles: Sequence[str],
    rows: Iterable[Sequence[Any]],
) -> int:
    """Создаёт новый файл с одним листом: шапка и строки.

    Нужно для выгрузок, у которых нет исходного файла-образца, — например
    таблицы оплат. Числа записываются числами, а не строками: иначе в Excel по
    колонке суммы не посчитать итог.
    """
    workbook = openpyxl.Workbook()
    try:
        worksheet = workbook.active
        # Excel не принимает в имени листа : \\ / ? * [ ] и больше 31 знака.
        worksheet.title = "".join(
            ch for ch in (sheet_name or "Лист1") if ch not in ':\\/?*[]')[:31] or "Лист1"
        worksheet.append(list(titles))
        for cell in worksheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        written = 0
        for row in rows:
            worksheet.append(list(row))
            written += 1
        worksheet.freeze_panes = "A2"
        for index, title in enumerate(titles, start=1):
            letter = openpyxl.utils.get_column_letter(index)
            worksheet.column_dimensions[letter].width = min(max(len(str(title)) + 4, 12), 46)
        worksheet.auto_filter.ref = worksheet.dimensions
        workbook.save(destination)
        return written
    finally:
        workbook.close()


def write_values(
    path: str,
    destination: str,
    updates: Iterable[tuple[int, int, Any]],
    overwrite: bool = False,
    sheet_name: str | None = None,
) -> int:
    """Записывает значения в копию файла. Формулы и оформление сохраняются.

    Файл открывается без data_only, иначе openpyxl заменил бы формулы их
    кэшированными значениями при сохранении.
    """
    workbook = open_workbook(path, data_only=False)
    try:
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.worksheets[0]
        written = 0
        for row, column, value in updates:
            if value is None:
                continue
            cell = worksheet.cell(row=row, column=column)
            if cell.value not in (None, "") and not overwrite:
                continue
            cell.value = value
            written += 1
        workbook.save(destination)
        return written
    finally:
        workbook.close()
