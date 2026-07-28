"""Тесты переноса заказа из выгрузки 1С в бланк поставщика."""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core import order


@pytest.fixture
def export(tmp_path: Path) -> str:
    """Выгрузка 1С: двухуровневая шапка и несколько колонок «Заказ»."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Производитель", "Артикул", "Номенклатура", "Штрихкод", "ИтогоЗаказ",
                  "Магазин А", None, "Магазин Б", None])
    sheet.append([None, None, None, None, None, "Остаток", "Заказ", "Остаток", "Заказ"])
    sheet.append(["БРЕНД", None, None, None, None, 10, None, 5, None])          # итоговая строка
    sheet.append(["БРЕНД", "ART-1", "Гель для душа (300мл)", "4600000000017", 12, 3, None, 1, None])
    sheet.append(["БРЕНД", "ART-2", "Крем для рук (50мл)", "4600000000024", 5, 2, None, 0, None])
    sheet.append(["БРЕНД", "ART-3", "Скраб для тела (200мл)", "4600000000031", 7, 1, None, 2, None])
    sheet.append(["БРЕНД", "ART-4", "Мыло твёрдое (75г)", "4600000000048", None, 4, None, 3, None])
    path = tmp_path / "выгрузка.xlsx"
    book.save(path)
    return str(path)


@pytest.fixture
def form(tmp_path: Path) -> str:
    """Бланк: шапка не в первой строке, штрихкод не под своим заголовком."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Дата"])
    sheet.append([None])
    sheet.append(["Штрихкод", None, "Артикул", "Наименование товара", "Цена", "Заказ\n(шт.)", "Сумма"])
    sheet.append(["Категория"])
    sheet.append([None, "4600000000017", "ART-1", "Гель для душа (300мл)", 100, None, "=F5*E5"])
    sheet.append([None, "4600000000024", "art-2", "Крем для рук (50мл)", 200, None, "=F6*E6"])
    sheet.append([None, "4600000000055", "ART-9", "Гель для душа ТЕСТЕР (300мл)", 0, None, None])
    sheet.append([None, "4600000000031", "ART-3t", "ТЕСТЕР Скраб для тела (200мл)", 0, None, None])
    path = tmp_path / "бланк.xlsx"
    book.save(path)
    return str(path)


def test_source_picks_filled_order_column(export: str) -> None:
    """Из нескольких колонок «Заказ» берётся та, где есть числа."""
    source = order.detect_source(export)
    assert source.title_of(source.quantity) == "ИтогоЗаказ"
    assert source.article is not None and source.title_of(source.article) == "Артикул"
    assert [o.title for o in source.options][0] == "ИтогоЗаказ"
    assert source.options[0].filled == 3


def test_target_finds_order_column_and_real_barcode(form: str) -> None:
    """Заголовок «Штрихкод» стоит над пустой колонкой — коды ищутся по данным."""
    target = order.detect_target(form)
    assert target.title_of(target.quantity) == "Заказ (шт.)"
    assert target.ean == 1  # колонка B, а не A под заголовком
    assert target.header_row == 2


def test_transfer_matches_by_article_ignoring_case(export: str, form: str) -> None:
    lines = order.transfer(order.detect_source(export), order.detect_target(form))
    by_article = {line.article: line for line in lines}
    assert by_article["ART-1"].target_row == 5
    assert by_article["ART-1"].method == "Артикул"
    assert by_article["ART-2"].target_row == 6  # в бланке «art-2»


def test_rows_without_order_are_skipped(export: str, form: str) -> None:
    """Позиции с пустым количеством в заказ не попадают."""
    lines = order.transfer(order.detect_source(export), order.detect_target(form))
    assert {line.article for line in lines} == {"ART-1", "ART-2", "ART-3"}


def test_tester_is_not_matched_automatically(export: str, form: str) -> None:
    """ART-3 есть в бланке только как ART-3t — привязка не делается сама."""
    lines = order.transfer(order.detect_source(export), order.detect_target(form))
    line = next(l for l in lines if l.article == "ART-3")
    assert not line.matched
    assert line.suggestions
    assert any(s.tester for s in line.suggestions)


def test_manual_assignment_and_updates(export: str, form: str) -> None:
    target = order.detect_target(form)
    lines = order.transfer(order.detect_source(export), target)
    line = next(l for l in lines if l.article == "ART-3")
    line.assign(line.suggestions[0].row)

    updates = order.build_updates(lines, target)
    assert (5, 6, 12) in updates          # ART-1 → строка 5, колонка F
    assert len(updates) == 3
    assert all(isinstance(value, int) for _, _, value in updates)


def test_summary_counts_units(export: str, form: str) -> None:
    lines = order.transfer(order.detect_source(export), order.detect_target(form))
    stats = order.summarize(lines)
    assert stats["перенесено"] == 2 and stats["штук перенесено"] == 17
    assert stats["не найдено"] == 1 and stats["штук потеряно"] == 7


@pytest.mark.parametrize(
    ("article", "name", "expected"),
    [("ART-3t", "Скраб", True), ("ART-3", "ТЕСТЕР Скраб", True), ("ART-3", "Скраб", False)],
)
def test_tester_detection(article: str, name: str, expected: bool) -> None:
    assert order.is_tester(article, name) is expected


# --- выбор листа --------------------------------------------------------------

@pytest.fixture
def two_sheets(tmp_path: Path) -> str:
    """Книга, где первым идёт лист, который не является бланком заказа."""
    book = openpyxl.Workbook()
    archive = book.active
    archive.title = "Архив"
    archive.append(["Артикул", "Наименование товара", "Цена", "Заказ, шт"])
    archive.append(["OLD-1", "Прошлогодняя позиция", 100, None])

    price = book.create_sheet("Прайс")
    price.append(["Артикул", "Наименование товара", "Цена", "ЗАКАЗ, шт"])
    price.append(["ART-1", "Гель для душа (300мл)", 100, None])
    path = tmp_path / "прайс.xlsx"
    book.save(path)
    return str(path)


def test_explicit_sheet_wins_over_detection(two_sheets: str) -> None:
    """Автоопределение берёт первый подходящий лист, выбор пользователя — важнее."""
    assert order.detect_target(two_sheets).sheet == "Архив"
    chosen = order.open_sheet(two_sheets, "Прайс", source=False)
    assert chosen.sheet == "Прайс"
    assert chosen.title_of(chosen.quantity) == "ЗАКАЗ, шт"


def test_open_sheet_without_choice_detects(two_sheets: str) -> None:
    assert order.open_sheet(two_sheets, None, source=False).sheet == "Архив"


# --- исключения ---------------------------------------------------------------

@pytest.fixture
def diffuser_export(tmp_path: Path) -> str:
    """Позиция, которой в бланке соответствует строка с другим названием."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Артикул", "Номенклатура", "Штрихкод", "ИтогоЗаказ"])
    sheet.append(["ZR-DIF-1", "Диффузор для ароматерапии Z&R Ветивер, Лимон", "4600000000109", 6])
    path = tmp_path / "выгрузка-диффузор.xlsx"
    book.save(path)
    return str(path)


@pytest.fixture
def diffuser_form(tmp_path: Path) -> str:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Артикул", "Наименование товара", "Цена", "Заказ (шт.)"])
    sheet.append(["SUP-77", "Мыло жидкое Дубовый мох (300мл)", 200, None])
    sheet.append(["SUP-90", "Диффузор. Средство для ароматизации помещений "
                            "Ветивер, Лимон (212,5мл)", 900, None])
    path = tmp_path / "бланк-диффузор.xlsx"
    book.save(path)
    return str(path)


def _diffuser_alias() -> order.AliasBook:
    return order.AliasBook([order.Alias(
        source_article="ZR-DIF-1",
        source_name="Диффузор для ароматерапии Z&R Ветивер, Лимон",
        target_article="SUP-90",
        target_name="Диффузор. Средство для ароматизации помещений Ветивер, Лимон (212,5мл)",
    )])


def test_without_alias_position_is_not_matched(diffuser_export: str, diffuser_form: str) -> None:
    lines = order.transfer(order.detect_source(diffuser_export), order.detect_target(diffuser_form))
    assert not lines[0].matched


def test_alias_binds_position_with_different_name(diffuser_export: str, diffuser_form: str) -> None:
    lines = order.transfer(
        order.detect_source(diffuser_export),
        order.detect_target(diffuser_form),
        _diffuser_alias(),
    )
    assert lines[0].target_row == 3
    assert lines[0].method == "Исключение"


def test_alias_finds_row_after_it_moved(tmp_path: Path, diffuser_export: str) -> None:
    """Бланк нового месяца: та же позиция стоит в другой строке."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Артикул", "Наименование товара", "Цена", "Заказ (шт.)"])
    for _ in range(4):
        sheet.append(["SUP-00", "Новинка", 100, None])
    sheet.append(["SUP-90", "Диффузор. Средство для ароматизации помещений "
                            "Ветивер, Лимон (212,5мл)", 900, None])
    path = tmp_path / "бланк-новый.xlsx"
    book.save(path)

    lines = order.transfer(
        order.detect_source(diffuser_export), order.detect_target(str(path)), _diffuser_alias())
    assert lines[0].target_row == 6


def test_remembering_alias_replaces_previous(diffuser_export: str, diffuser_form: str) -> None:
    target = order.detect_target(diffuser_form)
    index = order.TargetIndex(target)
    line = order.read_orders(order.detect_source(diffuser_export))[0]

    book = order.AliasBook()
    book.remember(line, index.info(2))
    book.remember(line, index.info(3))
    assert len(book) == 1
    assert book.find(line).target_article == "SUP-90"


def test_alias_is_restored_from_saved_data(diffuser_export: str, diffuser_form: str) -> None:
    saved = [alias.as_dict() for alias in _diffuser_alias().items]
    restored = order.AliasBook(order.Alias.from_dict(item) for item in saved)
    lines = order.transfer(
        order.detect_source(diffuser_export), order.detect_target(diffuser_form), restored)
    assert lines[0].method == "Исключение"


# --- поиск по бланку ----------------------------------------------------------

def test_search_finds_row_by_article_fragment(form: str) -> None:
    index = order.TargetIndex(order.detect_target(form))
    assert [s.row for s in index.search("art-2")] == [6]


def test_search_finds_row_by_words_in_any_order(diffuser_form: str) -> None:
    index = order.TargetIndex(order.detect_target(diffuser_form))
    found = index.search("лимон диффузор")
    assert found and found[0].row == 3


def test_search_by_barcode(form: str) -> None:
    index = order.TargetIndex(order.detect_target(form))
    assert [s.row for s in index.search("4600000000017")] == [5]


def test_search_tolerates_typo(diffuser_form: str) -> None:
    index = order.TargetIndex(order.detect_target(diffuser_form))
    found = index.search("ветиве лимн")
    assert found and found[0].row == 3


def test_empty_search_returns_nothing(form: str) -> None:
    index = order.TargetIndex(order.detect_target(form))
    assert index.search("   ") == []


# --- характеристика -----------------------------------------------------------

@pytest.fixture
def shades_export(tmp_path: Path) -> str:
    """Два оттенка одного товара: названия совпадают, различает характеристика."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Артикул", "Номенклатура", "Характеристика", "Штрихкод", "ИтогоЗаказ"])
    sheet.append(["ZR-01", "Тональный крем для лица", "20 мл, 01, светлый", "4600000000116", 5])
    sheet.append(["ZR-02", "Тональный крем для лица", "20 мл, 02, теплый", "4600000000123", 7])
    path = tmp_path / "оттенки.xlsx"
    book.save(path)
    return str(path)


@pytest.fixture
def marked_form(tmp_path: Path) -> str:
    """Бланк с колонкой пометок и словами-ловушками, которые пометками не являются."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["Артикул", "Наименование товара", "NEW/ХИТ", "Заказ (шт.)"])
    sheet.append(["SUP-01", "Тональный крем для лица 01 20 мл", "New", None])
    sheet.append(["SUP-02", "Тональный крем для лица 02 20 мл", "Хит", None])
    sheet.append(["SUP-03", "Пудра рассыпчатая", "Новый артикул", None])
    sheet.append(["SUP-04", "Кейс силиконовый под спонжи", None, None])
    sheet.append(["SUP-05", "Помада лимитированной серии", "Limited", None])
    path = tmp_path / "бланк-пометки.xlsx"
    book.save(path)
    return str(path)


def test_trait_is_read_from_export(shades_export: str) -> None:
    source = order.detect_source(shades_export)
    assert source.title_of(source.trait) == "Характеристика"
    lines = order.read_orders(source)
    assert [line.trait for line in lines] == ["20 мл, 01, светлый", "20 мл, 02, теплый"]
    assert lines[0].title == "Тональный крем для лица · 20 мл, 01, светлый"


def test_alias_for_one_shade_does_not_bind_another(shades_export: str, marked_form: str) -> None:
    """Название у оттенков общее — исключение не должно перетекать на соседний."""
    target = order.detect_target(marked_form)
    index = order.TargetIndex(target)
    first, second = order.read_orders(order.detect_source(shades_export))

    book = order.AliasBook()
    book.remember(first, index.info(2))
    book.remember(second, index.info(3))
    assert len(book) == 2

    lines = order.transfer(order.detect_source(shades_export), target, book)
    assert [line.target_row for line in lines] == [2, 3]
    assert all(line.method == "Исключение" for line in lines)


def test_alias_without_trait_still_works(shades_export: str, marked_form: str) -> None:
    """Исключения, сохранённые до появления характеристики, не теряются."""
    saved = order.Alias(source_article="ZR-01", source_name="Тональный крем для лица",
                        target_article="SUP-01", target_name="")
    lines = order.transfer(order.detect_source(shades_export),
                           order.detect_target(marked_form), order.AliasBook([saved]))
    assert lines[0].target_row == 2


# --- пометки поставщика -------------------------------------------------------

@pytest.mark.parametrize(
    ("text", "expected"),
    [
        ("New", ("NEW",)),
        ("NEW/ХИТ", ("NEW", "ХИТ")),
        ("Limited", ("LIMITED",)),
        ("Новинка", ("НОВИНКА",)),
        ("Новый артикул", ()),
        ("Кейс силиконовый под спонжи", ()),
        ("Вывод из ассортимента", ()),
    ],
)
def test_detect_marks(text: str, expected: tuple[str, ...]) -> None:
    assert order.detect_marks([text]) == expected


def test_marks_are_collected_per_row(marked_form: str) -> None:
    index = order.TargetIndex(order.detect_target(marked_form))
    marks = {entry.row: entry.marks for entry in index.entries if entry.marks}
    assert marks == {2: ("NEW",), 3: ("ХИТ",), 6: ("LIMITED",)}


def test_highlights_skip_rows_already_ordered(shades_export: str, marked_form: str) -> None:
    target = order.detect_target(marked_form)
    index = order.TargetIndex(target)
    lines = order.transfer(order.detect_source(shades_export), target,
                           order.AliasBook([order.Alias(
                               source_article="ZR-01", source_name="Тональный крем для лица",
                               source_trait="20 мл, 01, светлый", target_article="SUP-01")]))
    taken = {line.target_row for line in lines if line.matched}
    assert [entry.row for entry in index.highlights(taken)] == [3, 6]


def test_added_position_is_written_to_form(marked_form: str) -> None:
    target = order.detect_target(marked_form)
    index = order.TargetIndex(target)
    line = order.OrderLine.from_target(index.info(6), 4)

    assert line.added and line.matched
    assert line.method == "LIMITED"
    assert order.build_updates([line], target) == [(6, 4, 4)]
    assert order.summarize([line])["добавлено"] == 1
