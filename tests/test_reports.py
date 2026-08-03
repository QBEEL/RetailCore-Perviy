"""Отчётность: разбор выгрузок, правила магазинов, сводная и оформление файла."""
from __future__ import annotations

import openpyxl
import pytest

from app.core.reports import (
    Field,
    Metric,
    ReportProfile,
    SaleRow,
    StoreRule,
    apply_rules,
    build_map,
    default_profile,
    detect_mapping,
    normalize,
    period_of,
    read_source,
    store_from_name,
)
from app.core.reports import export, pivot, service, store


# --- вспомогательное --------------------------------------------------------------

HEADER = ["Магазин", None, None, "Номенклатура с характеристикой", None,
          "Категория товара", "Бренд", "Акция", "% авт.", "Год", "Месяц",
          "Количество", "Сумма авт.", "Скидка бонусами", "Сумма",
          "Начислено бонусов"]


def _row(store_name, item, promo, qty, discount, *, month=6, year=2026,
         brand="BIOREPAIR", bonus=0.0):
    return [store_name, None, None, item, None, "Личная гигиена", brand, promo,
            "30", year, month, qty, discount, bonus, 0, 0]


def _workbook(tmp_path, rows, *, name="выгрузка.xlsx", header=HEADER,
              preamble=True, total=True):
    """Книга в том виде, в каком её отдаёт 1С: параметры сверху, «Итого» снизу."""
    book = openpyxl.Workbook()
    sheet = book.active
    if preamble:
        sheet.append([])
        sheet.append(["Параметры:", None, "Период: 01.06.2026 - 30.06.2026"])
        sheet.append([None, None, "Подразделение: Уссурийск; Универмаг"])
        sheet.append(["Отбор:", None, "Номенклатура В группе из списка"])
        sheet.append([])
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    if total:
        sheet.append(["Итого", None, None, None, None, None, None, None, None,
                      None, None, sum(r[11] for r in rows),
                      sum(r[12] for r in rows), 0, 0, 0])
    path = tmp_path / name
    book.save(path)
    return str(path)


# --- разбор исходников ---------------------------------------------------------------

def test_header_found_below_report_parameters(tmp_path):
    """Шапка ищется ниже блока «Параметры» — там встречаются те же слова."""
    path = _workbook(tmp_path, [_row("Универмаг", "Паста", "Акция 30%", 2, 100.0)])
    source = read_source(path)

    assert source.mapping.header_row == 5
    assert source.mapping.column_of(Field.ITEM) == 3
    assert source.mapping.column_of(Field.QTY) == 11
    assert source.mapping.usable


def test_total_row_is_not_a_product(tmp_path):
    """Строка «Итого» из выгрузки не должна стать позицией отчёта."""
    path = _workbook(tmp_path, [_row("Универмаг", "Паста", "Акция", 2, 100.0)])
    source = read_source(path)

    assert len(source.rows) == 1
    assert source.rows[0].item == "Паста"


def test_revenue_and_auto_discount_are_told_apart(tmp_path):
    """«Сумма» и «Сумма авт.» — разные колонки, и путать их нельзя.

    Именно на этой паре строится отчёт: скидка идёт из «Сумма авт.», а не из
    выручки, и ошибка здесь тихо утроила бы цифры поставщику.
    """
    mapping = detect_mapping([HEADER])

    assert mapping.column_of(Field.AUTO_DISCOUNT) == 12
    assert mapping.column_of(Field.REVENUE) == 14


def test_numbers_arrive_as_text(tmp_path):
    """1С выгружает числа строкой с пробелами и запятой чаще, чем хотелось бы."""
    rows = [_row("Универмаг", "Паста", "Акция", "2", "1 049,99")]
    path = _workbook(tmp_path, rows, total=False)
    source = read_source(path)

    assert source.rows[0].qty == 2
    assert source.rows[0].auto_discount == pytest.approx(1049.99)


def test_store_taken_from_file_name_when_column_missing(tmp_path):
    """Точечная выгрузка одного магазина: колонки «Магазин» в ней нет."""
    header = [h for h in HEADER if h != "Магазин"]
    rows = [_row("", "Паста", "Акция", 1, 10.0)[1:]]
    path = _workbook(tmp_path, rows, header=header,
                     name="Продажи Уссурийск июнь 2026.xlsx", total=False)
    source = read_source(path)

    assert not source.mapping.has(Field.STORE)
    assert source.rows[0].store == "Уссурийск"


def test_store_from_name_drops_month_and_noise():
    assert store_from_name("/x/Продажи Уссурийск июнь 2026.xlsx") == "Уссурийск"
    assert store_from_name("/x/Калина Молл.xlsx") == "Калина Молл"


def test_explicit_store_hint_wins_over_file_name(tmp_path):
    header = [h for h in HEADER if h != "Магазин"]
    rows = [_row("", "Паста", "Акция", 1, 10.0)[1:]]
    path = _workbook(tmp_path, rows, header=header, name="выгрузка_7.xlsx",
                     total=False)
    source = read_source(path, store_hint="Седанка Сити")

    assert source.rows[0].store == "Седанка Сити"


# --- правила объединения магазинов ------------------------------------------------------

def test_chain_of_rules_is_flattened():
    """A → B → C должно довезти продажи A сразу до C."""
    mapping = build_map([StoreRule(source="A", target="B"),
                         StoreRule(source="B", target="C")])

    assert mapping.valid
    assert mapping.resolve("A") == "C"
    assert mapping.resolve("B") == "C"


def test_cycle_disables_all_transfers():
    """Кольцо не применяется частично: половина переноса хуже, чем ни одного."""
    mapping = build_map([StoreRule(source="A", target="B"),
                         StoreRule(source="B", target="A")])

    assert not mapping.valid
    assert mapping.cycles
    assert mapping.resolve("A") == "A"


def test_disabled_rule_is_ignored():
    mapping = build_map([StoreRule(source="A", target="B", enabled=False)])

    assert mapping.resolve("A") == "A"


def test_store_names_compared_loosely():
    """Регистр, двойные пробелы и «ё» в выгрузках гуляют."""
    mapping = build_map([StoreRule(source="сахалин  ПП", target="Универмаг")])

    assert mapping.resolve("Сахалин ПП") == "Универмаг"
    assert normalize("Артём") == normalize("Артем")


def test_substring_is_not_a_match():
    """«Артем (Первый Парфюмерный)» — не «Первый Парфюмерный»."""
    mapping = build_map([StoreRule(source="Первый Парфюмерный", target="Универмаг")])

    assert mapping.resolve("Артем (Первый Парфюмерный)") == "Артем (Первый Парфюмерный)"
    assert mapping.resolve("Первый Парфюмерный") == "Универмаг"


def test_transfer_keeps_original_store_name():
    """Исходный магазин сохраняется: без него перенос нечем объяснить."""
    rows = [SaleRow(store="Интернет-магазин", item="Паста", qty=1)]
    moved, _, count = apply_rules(
        rows, [StoreRule(source="Интернет-магазин", target="Универмаг")])

    assert count == 1
    assert moved[0].store == "Универмаг"
    assert moved[0].source_store == "Интернет-магазин"


# --- сводная -----------------------------------------------------------------------------

def _sales():
    return [
        SaleRow(store="Универмаг", item="Паста A", promo="Скидка 20%", qty=2,
                auto_discount=100.0, year=2026, month=6),
        SaleRow(store="Универмаг", item="Паста A", promo="Скидка 20%", qty=3,
                auto_discount=150.0, year=2026, month=6),
        SaleRow(store="Уссурийск", item="Паста B", promo="Скидка 30%", qty=1,
                auto_discount=90.0, year=2026, month=6),
        SaleRow(store="Уссурийск", item="Паста C", promo="", qty=7,
                auto_discount=0.0, bonus_discount=500.0, year=2026, month=6),
    ]


def test_rows_without_promo_are_dropped():
    """То самое «удаление лишнего», которое делалось руками."""
    table, _ = pivot.build(_sales(), default_profile("П"), [])

    assert table.used_rows == 3
    assert table.dropped_rows == 1
    assert [row.keys[0] for row in table.rows] == ["Паста A", "Паста B"]


def test_same_product_is_summed_across_stores():
    table, _ = pivot.build(_sales(), default_profile("П"), [])
    first = table.rows[0]

    assert first.cells[0].value == 5
    assert first.cells[1].value == pytest.approx(250.0)


def test_totals_match_the_sum_of_rows():
    table, _ = pivot.build(_sales(), default_profile("П"), [])

    for position, total in enumerate(table.totals):
        assert total.value == pytest.approx(
            sum(row.cells[position].value for row in table.rows))


def test_period_comes_from_the_data_not_today():
    table, _ = pivot.build(_sales(), default_profile("П"), [])

    assert table.period.title == "Июнь 2026"
    assert table.file_name() == "Июнь 2026"


def test_stray_month_does_not_rename_the_report():
    """Одна поздняя проводка соседнего месяца не должна назвать файл её именем."""
    rows = _sales()
    rows.append(SaleRow(store="Универмаг", item="Паста A", promo="Скидка 20%",
                        qty=1, auto_discount=10.0, year=2026, month=7))

    assert period_of(rows).title == "Июнь 2026"


def test_promo_columns_sorted_naturally():
    """«20%» должен идти перед «100%», а не после — как при обычной сортировке."""
    rows = [
        SaleRow(item="Паста", promo="Список_100%", qty=1, auto_discount=1, month=6, year=2026),
        SaleRow(item="Паста", promo="Список_20%", qty=1, auto_discount=1, month=6, year=2026),
    ]
    table, _ = pivot.build(rows, default_profile("П"), [])

    assert [group.label for group in table.groups] == ["Список_20%", "Список_100%"]


def test_all_sales_kept_when_filter_switched_off():
    profile = default_profile("П")
    profile.filters.promo_only = False
    table, _ = pivot.build(_sales(), profile, [])

    assert table.dropped_rows == 0
    assert len(table.groups) == 3  # включая пустую акцию


def test_total_discount_metric_adds_both_kinds():
    profile = default_profile("П")
    profile.filters.promo_only = False
    profile.metrics = [Metric.TOTAL_DISCOUNT]
    table, _ = pivot.build(_sales(), profile, [])

    assert sum(cell.value for cell in table.totals) == pytest.approx(840.0)


def test_store_rules_apply_before_the_pivot():
    profile = default_profile("П")
    table, _ = pivot.build(_sales(), profile,
                           [StoreRule(source="Уссурийск", target="Универмаг")])

    assert table.moved_rows == 2
    assert table.stores == ["Универмаг"]


# --- оформление файла ----------------------------------------------------------------------

def _built(tmp_path, profile=None):
    table, _ = pivot.build(_sales(), profile or default_profile("SmartBeauty"), [])
    path = str(tmp_path / "отчёт.xlsx")
    export.save(table, path, author="Иванов Е.")
    return table, openpyxl.load_workbook(path)


def test_exported_file_has_a_header_and_a_frozen_table(tmp_path):
    table, book = _built(tmp_path)
    sheet = book.active

    assert sheet.title == "Июнь 2026"
    assert "SmartBeauty" in str(sheet["A1"].value)
    assert "Период: 01.06.2026" in str(sheet["A2"].value)
    assert sheet.freeze_panes is not None


def test_numbers_are_rounded_and_formatted(tmp_path):
    """В прежнем отчёте в ячейке лежало «3160.1000000000004» без формата."""
    rows = [SaleRow(item="Паста", promo="Акция", qty=1, auto_discount=3160.1000000000004,
                    year=2026, month=6)]
    table, _ = pivot.build(rows, default_profile("П"), [])
    path = str(tmp_path / "числа.xlsx")
    export.save(table, path)
    sheet = openpyxl.load_workbook(path).active

    money = [cell for row in sheet.iter_rows() for cell in row
             if isinstance(cell.value, float)]
    assert money and all(cell.number_format == "#,##0.00" for cell in money)
    assert money[0].value == pytest.approx(3160.10)


def test_autofilter_excludes_the_totals_row(tmp_path):
    """Попав в диапазон фильтра, «Итого» уезжает вместе с отфильтрованными строками."""
    table, book = _built(tmp_path)
    sheet = book.active
    last = int(sheet.auto_filter.ref.split(":")[1].lstrip("ABCDEFGHIJKLMNOP"))

    total_row = next(row for row in sheet.iter_rows()
                     if row[0].value == "Итого")
    assert total_row[0].row > last


def test_column_widths_are_bounded(tmp_path):
    """Название товара длиной в полторы сотни знаков не должно растянуть колонку."""
    long_name = "Зубная паста " + "очень длинное название " * 8
    rows = [SaleRow(item=long_name, promo="Акция", qty=1, auto_discount=1.0,
                    year=2026, month=6)]
    table, _ = pivot.build(rows, default_profile("П"), [])
    path = str(tmp_path / "ширина.xlsx")
    export.save(table, path)
    sheet = openpyxl.load_workbook(path).active

    assert sheet.column_dimensions["A"].width <= export.MAX_WIDTH


def test_stores_sheet_added_on_demand(tmp_path):
    profile = default_profile("П")
    profile.stores_sheet = True
    _, book = _built(tmp_path, profile)

    assert "По магазинам" in book.sheetnames
    assert book["По магазинам"]["A4"].value in ("Универмаг", "Уссурийск")


def test_file_name_follows_the_template(tmp_path):
    profile = default_profile("SmartBeauty")
    profile.file_name = "Отчёт {Поставщик} {Месяц} {Год}"
    table, _ = pivot.build(_sales(), profile, [])

    assert table.file_name() == "Отчёт SmartBeauty Июнь 2026"
    assert export.default_name(table, "/tmp").endswith(
        "Отчёт SmartBeauty Июнь 2026.xlsx")


# --- профиль -------------------------------------------------------------------------------

def test_profile_survives_a_round_trip():
    profile = default_profile("SmartBeauty")
    profile.columns = [Field.STORE, Field.PROMO]
    profile.metrics = [Metric.QTY, Metric.TOTAL_DISCOUNT]
    profile.filters.brands = ["BLANX"]
    restored = ReportProfile.from_dict(profile.as_dict())

    assert restored.columns == [Field.STORE, Field.PROMO]
    assert restored.metrics == [Metric.QTY, Metric.TOTAL_DISCOUNT]
    assert restored.filters.brands == ["BLANX"]


def test_unknown_field_in_stored_profile_is_skipped():
    """Профиль мог быть записан более новой версией приложения."""
    profile = ReportProfile.from_dict({"name": "x", "rows": ["item", "выдумка"]})

    assert profile.rows == [Field.ITEM]


def test_profile_filled_with_plain_strings_still_builds():
    """Роли могли прийти строками, пройдя через чужой контейнер.

    `Field` и `Metric` наследуют `str`, и Qt возвращает их из хранилища
    элемента списка обычными строками. Список при этом выглядит рабочим —
    сравнение с ролью проходит, подпись находится, — и падает лишь там, где
    нужен сам объект. Отчёт обязан собраться и в таком виде.
    """
    profile = default_profile("П")
    profile.rows = ["item"]
    profile.columns = ["promo"]
    profile.metrics = ["qty", "auto_discount"]

    table, _ = pivot.build(_sales(), profile, [])

    assert table.row_fields == [Field.ITEM]
    assert table.metrics == [Metric.QTY, Metric.AUTO_DISCOUNT]
    assert table.rows[0].cells[0].value == 5


def test_profile_with_plain_strings_can_be_saved(tmp_path):
    """Сохранение такого профиля тоже не должно падать."""
    profile = default_profile("П")
    profile.name = "П"
    profile.metrics = ["qty"]

    saved = store.save_profile(profile, str(tmp_path / "reports.db"))

    assert saved.as_dict()["metrics"] == ["qty"]


def test_metrics_survive_the_profile_dialog():
    """Диалог обязан вернуть перечисления, а не строки.

    Это и была причина ошибки «'str' object has no attribute 'value'»: первый
    отчёт собирался из профиля, прочитанного из базы, а после открытия диалога
    роли в памяти подменялись строками, и вторая сборка падала.
    """
    pytest.importorskip("PySide6")
    import os

    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication

    from app.ui.widgets.report_dialogs import ProfileDialog

    application = QApplication.instance() or QApplication([])
    profile = default_profile("SmartBeauty")
    profile.name = "SmartBeauty"
    dialog = ProfileDialog(profile)
    dialog._apply()

    assert all(isinstance(role, Field) for role in profile.rows)
    assert all(isinstance(role, Field) for role in profile.columns)
    assert all(isinstance(metric, Metric) for metric in profile.metrics)
    assert profile.metrics == [Metric.QTY, Metric.AUTO_DISCOUNT]
    dialog.deleteLater()


# --- хранилище ------------------------------------------------------------------------------

def test_profiles_are_stored_and_read_back(tmp_path):
    path = str(tmp_path / "reports.db")
    profile = default_profile("SmartBeauty")
    profile.name = "SmartBeauty — акции"
    saved = store.save_profile(profile, path)

    restored = store.list_profiles(path)
    assert len(restored) == 1
    assert restored[0].id == saved.id
    assert restored[0].metrics == profile.metrics


def test_rule_source_is_unique(tmp_path):
    """Один магазин нельзя отправить сразу в два — это молча удвоенные продажи."""
    path = str(tmp_path / "reports.db")
    store.save_rule(StoreRule(source="Интернет-магазин", target="Универмаг"), path)
    store.save_rule(StoreRule(source="интернет-магазин", target="Уссурийск"), path)

    rules = store.list_rules(path)
    assert len(rules) == 1
    assert rules[0].target == "Уссурийск"


def test_rule_into_itself_is_refused(tmp_path):
    path = str(tmp_path / "reports.db")
    with pytest.raises(ValueError):
        store.save_rule(StoreRule(source="Универмаг", target="универмаг"), path)


# --- сценарий целиком -------------------------------------------------------------------------

def test_two_files_merge_into_one_report(tmp_path):
    """Разные менеджеры прислали свои магазины — отчёт должен быть общим."""
    first = _workbook(tmp_path, [_row("Универмаг", "Паста A", "Акция 30%", 2, 100.0)],
                      name="универмаг.xlsx")
    second = _workbook(tmp_path, [_row("Уссурийск", "Паста A", "Акция 30%", 3, 150.0)],
                       name="уссурийск.xlsx")

    result = service.build_from_paths([first, second], default_profile("П"))

    assert result.ok
    assert result.table.used_rows == 2
    assert len(result.table.rows) == 1  # один товар, продажи сложились
    assert result.table.rows[0].cells[0].value == 5
    assert result.table.rows[0].cells[1].value == pytest.approx(250.0)


def test_unreadable_file_does_not_sink_the_others(tmp_path):
    good = _workbook(tmp_path, [_row("Универмаг", "Паста", "Акция", 1, 10.0)])
    broken = tmp_path / "битый.xlsx"
    broken.write_text("это не книга Excel", encoding="utf-8")

    result = service.build_from_paths([good, str(broken)], default_profile("П"))

    assert result.ok
    assert any("битый.xlsx" in warning for warning in result.warnings)


def test_empty_result_explains_itself(tmp_path):
    """Все строки отсеялись фильтром — пользователь должен узнать почему."""
    path = _workbook(tmp_path, [_row("Универмаг", "Паста", "", 1, 10.0)])

    result = service.build_from_paths([path], default_profile("П"))

    assert not result.ok
    assert any("фильтр" in warning for warning in result.warnings)
