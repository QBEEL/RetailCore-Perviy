"""Тесты переоценки: разбор артикула, чтение шаблона 1С, сравнение, выгрузка."""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.article import (
    DEFAULT_MODIFIER_SEPARATORS,
    article_variants,
    parse_article,
    split_articles,
)
from app.core.normalize import KeyOptions, code_key
from app.core.pricing import (
    MatchOptions,
    PriceStatus,
    SupplierProfile,
    SupplierProfiles,
    compare,
    export,
    load_supplier,
    load_template,
    match_lines,
    prepare_profile,
    suggest_price_map,
    suggest_profile_name,
)
from app.core.workbook import open_workbook


# --- разбор артикула ---------------------------------------------------------

def test_несколько_артикулов_в_одной_ячейке():
    """Ячейка 1С перечисляет варианты одной номенклатуры — каждый самостоятелен."""
    assert split_articles("zrp0050perG32/zrp0010perG32") == [
        "zrp0050perG32", "zrp0010perG32"]
    # Обратный слеш встречается в тех же выгрузках наравне с прямым.
    assert split_articles("zrp0010perG19\\zrp0050perG19") == [
        "zrp0010perG19", "zrp0050perG19"]
    assert len(split_articles("zrn0085difG08/zrn0212difG08/zrn0425difG08/zrn0850difG08")) == 4


def test_короткий_хвост_это_модификация_а_не_артикул():
    """«ABC123/30» — один товар с объёмом, а не два разных артикула."""
    assert split_articles("ABC123/30") == ["ABC123"]
    parsed = parse_article("ABC123/30")
    assert parsed.base_key == "abc123"
    assert parsed.modifier == "30"


def test_модификация_с_единицей_даёт_объём():
    parsed = parse_article("Cream-01/50ml")
    assert parsed.base_key == code_key("Cream-01")
    assert parsed.modifier == "50ml"
    assert parsed.quantity is not None
    assert parsed.quantity.value == 50


def test_длинный_хвост_модификацией_не_считается():
    """Иначе два разных артикула склеились бы в одну базу и цена ушла бы не туда."""
    parsed = parse_article("zrp0050perG32/zrp0010perG32")
    assert parsed.modifier == ""
    assert parsed.base_key == parsed.key


def test_единственный_короткий_артикул_не_отбрасывается():
    assert split_articles("1234") == ["1234"]


def test_варианты_содержат_и_части_и_базу():
    variants = article_variants("ABC123/50", modifier_separators=DEFAULT_MODIFIER_SEPARATORS)
    assert {v.base_key for v in variants} == {"abc123"}
    assert any(v.modifier == "50" for v in variants)


def test_разделители_задаются_вызывающей_стороной():
    assert split_articles("A-1|B-2", separators="|") == ["A-1", "B-2"]
    assert split_articles("A-1|B-2", separators="/") == ["A-1|B-2"]


def test_ключ_учитывает_настройки_строгости():
    assert code_key("ZRP-0010") == "zrp0010"
    assert code_key("ZRP-0010", KeyOptions(ignore_case=False)) == "ZRP0010"
    assert code_key("ZRP-0010", KeyOptions(ignore_symbols=False)) == "zrp-0010"


# --- подбор соответствия колонок ---------------------------------------------

class _Column:
    def __init__(self, title: str, priced: bool = True) -> None:
        self.title, self.priced, self.index, self.filled = title, priced, 0, 10


def test_новая_цена_выбирается_по_назначению_колонки():
    """Закупочная берёт закупочную колонку, розничная — розничную."""
    from app.core.pricing import PriceType

    types = [
        PriceType(name="Закупочная", old_column=6, price_column=9),
        PriceType(name="Прайс-лист Сахалин", old_column=12, price_column=15),
    ]
    columns = [
        _Column("РРЦ текущая"), _Column("РРЦ с 1/08/26"),
        _Column("Закупка старая"), _Column("Закупка нвоая"),
    ]
    mapping = suggest_price_map(types, columns)
    # «нвоая» — опечатка поставщика: слова «новая» в заголовке нет, и колонка
    # выбирается тем, что она единственная не помеченная как старая.
    assert mapping["Закупочная"] == "Закупка нвоая"
    assert mapping["Прайс-лист Сахалин"] == "РРЦ с 1/08/26"


def test_имя_профиля_без_служебных_слов_и_дат():
    assert suggest_profile_name("Переоценка Zielinski.xlsx") == "Zielinski"


def test_профиль_находится_по_имени_файла():
    profiles = SupplierProfiles([SupplierProfile(name="Zielinski")])
    assert profiles.for_file("C:/цены/Переоценка Zielinski 2026.xlsx") is not None
    assert profiles.for_file("C:/цены/Прайс Другого.xlsx") is None


# --- работа с настоящими файлами ---------------------------------------------

DATA = Path(__file__).resolve().parents[1] / "Для переоценки"
TEMPLATE = DATA / "шаблон выгрузки из 1С.xls"
SUPPLIER = DATA / "Переоценка Zielinski.xlsx"

needs_files = pytest.mark.skipif(
    not (TEMPLATE.exists() and SUPPLIER.exists()),
    reason="нет примеров файлов переоценки")


@pytest.fixture(scope="module")
def template():
    return load_template(str(TEMPLATE))


@pytest.fixture(scope="module")
def supplier():
    return load_supplier(str(SUPPLIER))


@needs_files
def test_шаблон_с_расширением_xls_читается(template):
    """1С выгружает xlsx под именем .xls — по расширению openpyxl его не откроет."""
    assert template.sheet_name == "Прайс-лист"
    assert len(template.records) > 300


@needs_files
def test_виды_цен_берутся_со_служебного_листа(template):
    names = [t.name for t in template.valid_types]
    assert names == ["Закупочная", "Прайс-лист Сахалин", "Прайс-лист Универмаг"]
    assert [t.price_column for t in template.valid_types] == [9, 15, 21]
    assert [t.old_column for t in template.valid_types] == [6, 12, 18]


@needs_files
def test_названием_товара_становится_товар_а_не_guid(template):
    """Рядом стоит «Уникальный идентификатор (Номенклатура)» — из GUID не взять объём."""
    assert template.article_column == 2
    assert template.name_column == 5
    assert any(record.quantity is not None for record in template.records)


@needs_files
def test_строки_разделители_прайса_пропускаются(supplier):
    """У строк с одной лишь категорией нет ни артикула, ни кода — это не товар."""
    from app.core.models import FieldRole

    identifiers = (FieldRole.ARTICLE, FieldRole.SKU, FieldRole.EAN)
    assert all(
        any(record.by_role.get(role) for role in identifiers)
        for record in supplier.records
    )
    assert len(supplier.records) > 700
    assert [c.title for c in supplier.price_columns]


@needs_files
def test_объём_берётся_из_категории_если_его_нет_в_названии(supplier):
    assert sum(1 for r in supplier.records if r.quantity is not None) > len(supplier.records) * 0.9


@needs_files
def test_дубли_артикула_разводятся_объёмом(template, supplier):
    """Одна ячейка артикула стоит в строках на 10 и 50 мл — цена не должна смешаться."""
    profile = prepare_profile(template, supplier, SupplierProfiles())
    lines = match_lines(template, supplier, MatchOptions())
    by_row = {line.row: line for line in lines}

    for row, expected in ((4, "zrp0050perG32"), (5, "zrp0010perG32")):
        line = by_row[row]
        assert line.matched, f"строка {row} осталась без варианта"
        assert line.supplier_article == expected
        assert "50 мл" in line.name or "10 мл" in line.name

    compare(lines, template, supplier, profile)
    assert by_row[4].status is PriceStatus.CHANGED


@needs_files
def test_доля_автоматического_сопоставления(template, supplier):
    profile = prepare_profile(template, supplier, SupplierProfiles())
    lines = match_lines(template, supplier, MatchOptions())
    stats = compare(lines, template, supplier, profile)
    assert stats.total == len(template.records)
    assert stats.rate > 85, f"сопоставилось только {stats.rate:.1f} %"
    assert stats.changed > 0
    assert stats.found == stats.changed + stats.unchanged + stats.no_price


@needs_files
def test_нечёткое_совпадение_само_не_применяется(template, supplier):
    """Цена соседнего аромата хуже, чем отсутствие цены."""
    lines = match_lines(template, supplier, MatchOptions())
    assert all(
        line.candidate.stage in ("Артикул", "EAN")
        for line in lines if line.matched)


@needs_files
def test_выгрузка_меняет_только_колонки_цены(template, supplier, tmp_path):
    profile = prepare_profile(template, supplier, SupplierProfiles())
    lines = match_lines(template, supplier, MatchOptions())
    compare(lines, template, supplier, profile)

    destination = tmp_path / "результат.xlsx"
    report = export(template, lines, str(destination))
    assert report.cells > 0

    before = open_workbook(str(TEMPLATE), data_only=False)
    after = openpyxl.load_workbook(destination)
    try:
        assert before.sheetnames == after.sheetnames
        old, new = before["Прайс-лист"], after["Прайс-лист"]
        assert (old.max_row, old.max_column) == (new.max_row, new.max_column)
        assert sorted(map(str, old.merged_cells.ranges)) == sorted(map(str, new.merged_cells.ranges))
        assert {k: v.width for k, v in old.column_dimensions.items()} == \
               {k: v.width for k, v in new.column_dimensions.items()}
        assert {k: v.hidden for k, v in old.column_dimensions.items()} == \
               {k: v.hidden for k, v in new.column_dimensions.items()}
        assert before.calculation.refMode == after.calculation.refMode == "R1C1"

        changed = set()
        for row in range(1, old.max_row + 1):
            for column in range(1, old.max_column + 1):
                a, b = old.cell(row, column), new.cell(row, column)
                assert a.number_format == b.number_format
                if a.value != b.value:
                    changed.add(column)
        assert changed == {9, 15, 21}, "изменились колонки помимо «Цена»"
        # Формула процента должна уцелеть: 1С пересчитает её при открытии.
        assert str(new["H3"].value).startswith("=IF(")
    finally:
        before.close()


@needs_files
def test_исключение_неизменных_строк_и_нумерация(template, supplier, tmp_path):
    profile = prepare_profile(template, supplier, SupplierProfiles())
    lines = match_lines(template, supplier, MatchOptions())
    stats = compare(lines, template, supplier, profile)

    destination = tmp_path / "только изменённые.xlsx"
    report = export(template, lines, str(destination), skip_unchanged=True)
    assert report.removed == stats.total - stats.changed

    sheet = openpyxl.load_workbook(destination)["Прайс-лист"]
    assert sheet.max_row == template.header_row + stats.changed
    assert [sheet.cell(row, 1).value for row in (3, 4, 5)] == [1, 2, 3]
    assert sheet.cell(sheet.max_row, 1).value == stats.changed


@needs_files
def test_ручная_привязка_пересчитывает_цену(template, supplier):
    profile = prepare_profile(template, supplier, SupplierProfiles())
    lines = match_lines(template, supplier, MatchOptions())
    compare(lines, template, supplier, profile)

    line = next(l for l in lines if l.status is PriceStatus.REVIEW and l.alternatives)
    line.assign(line.alternatives[0])
    assert line.manual and line.method == "Вручную"

    compare(lines, template, supplier, profile)
    assert line.status in (PriceStatus.CHANGED, PriceStatus.UNCHANGED, PriceStatus.NO_PRICE)

    line.clear()
    assert not line.matched and not line.cells


@needs_files
def test_старый_формат_xls_объясняется_понятно(tmp_path):
    fake = tmp_path / "старый.xls"
    fake.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64)
    with pytest.raises(ValueError, match="старом формате"):
        open_workbook(str(fake))
