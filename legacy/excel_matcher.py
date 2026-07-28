"""
Excel Data Matcher - standalone GUI app
"""
import subprocess, sys
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re, threading, os

BG = "#1a1d2e"; BG2 = "#252840"; BG3 = "#2e3250"
ACCENT = "#00d4aa"; ACCENT2 = "#7c3aed"
TEXT = "#e2e8f0"; TEXT_MUTE = "#8892a4"
RED = "#ef4444"; ORANGE = "#f97316"; GREEN = "#22c55e"
FN = ("Segoe UI", 10); FB = ("Segoe UI", 10, "bold")
FT = ("Segoe UI", 13, "bold"); FS = ("Segoe UI", 9)
FM = ("Consolas", 9)

STOPWORDS = {"sredstvo","dlya","pomeshcheniy","v","banke","kontsentrir","pavlik"}

def norm_vol(v):
    if not v: return ""
    s = str(v).lower().replace("yo","e").split(",")[0].strip()
    s = re.sub(r"\s+", "", s).replace("gr","g")
    s = re.sub(r"(\d+)[.,]\d+", lambda m: m.group(1), s)
    return s

def norm_name(n):
    if not n: return ""
    s = str(n).lower().replace("yo","e")
    s = re.sub(r"z\s*&\s*r\s*", "", s, flags=re.I)
    s = re.sub(r"\(\s*[\d,\.]+\s*[mgMG][lLgG]?[rR]?[^)]*\)", "", s)
    s = re.sub(r"\s*//?\s*\w+\s*$", "", s)
    s = re.sub(r",\s*", ", ", s)
    return re.sub(r"\s+", " ", s).strip()

def kwords(n):
    if not n: return frozenset()
    s = re.sub(r"z\s*&\s*r\s*","",str(n),flags=re.I).lower()
    s = re.sub(r"\(.*?\)","",s)
    ws = re.findall(r"[\w]+", s)
    return frozenset(w for w in ws if len(w)>2)

def jaccard(a,b):
    if not a or not b: return 0
    return len(a&b)/len(a|b)

def load_source(path, ce=0, ca=1, cr=3, crrc=5, skip=1):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    am,em,nm,nom,wm = {},{},{},{},{}
    rows=[]
    for i,row in enumerate(ws.iter_rows(values_only=True)):
        if i<skip: continue
        def g(idx): return row[idx] if idx<len(row) else None
        ean=str(g(ce)).strip() if g(ce) else None
        art=str(g(ca)).strip() if g(ca) else None
        ru=g(cr); rrc=g(crrc)
        if not ean and not art: continue
        e={"ean":ean,"article":art,"ru_name":ru,"rrc":rrc}
        rows.append(e)
        if art: am[art.lower()]=e
        if ean: em[ean]=e
        if ru:
            vm=re.search(r"\(\s*([\d,\.]+\s*[mgMGmlML][lLgGrR]?[^)]*)\)",str(ru))
            vol=norm_vol(vm.group(1)) if vm else ""
            nn=norm_name(str(ru)); kw=kwords(str(ru))
            nm.setdefault((nn,vol),e); nom.setdefault(nn,[]).append(e)
            wm.setdefault((kw,vol),[]).append(e)
    wb.close()
    return rows,am,em,nm,nom,wm

def load_target(path, ce=1, ca=2, cn=3, cv=4, crrc=5, ca2=10, skip=1):
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]
    rows=[]
    for i in range(skip+1, ws.max_row+1):
        def c(col): return ws.cell(row=i,column=col+1).value
        name=c(cn)
        if not name: continue
        rows.append({"row_idx":i,"ean":c(ce),"article":c(ca),"name":str(name),
                     "vol":c(cv),"rrc":c(crrc),"art2":c(ca2)})
    wb.close()
    return rows

def auto_match(tr, am, em, nm, nom, wm):
    found=None; reason=""
    art2=tr.get("art2"); ean=tr.get("ean"); art=tr.get("article")
    name=tr.get("name",""); vol=tr.get("vol")
    if art2 and isinstance(art2,str) and not art2.startswith("="):
        for a in [x.strip().lower() for x in art2.split("/")]:
            if a in am: found=am[a]; reason=f"Artikul(K):{a}"; break
    if not found and ean: found=em.get(str(ean).strip()); reason=f"EAN:{ean}"
    if not found and art: found=am.get(str(art).strip().lower()); reason=f"Artikul(C):{art}"
    if not found:
        nn=norm_name(name); nv=norm_vol(str(vol)) if vol else ""
        found=nm.get((nn,nv)); reason="Nazv+Vol"
    if not found:
        kw=kwords(name); nv=norm_vol(str(vol)) if vol else ""
        cands=wm.get((kw,nv),[])
        if cands: found=cands[0]; reason="Keywords+Vol"
    if not found:
        kw=kwords(name); nv=norm_vol(str(vol)) if vol else ""
        best=None; bsc=0
        for (skw,sv),entries in wm.items():
            sc=jaccard(kw,skw)
            if sv==nv: sc+=0.3
            if sc>bsc and sc>=0.72: bsc=sc; best=entries[0]
        if best: found=best; reason=f"Fuzzy({bsc:.0%})"
    return found, reason

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Data Matcher"); self.geometry("1350x820")
        self.configure(bg=BG); self.resizable(True,True)
        self.src_path=tk.StringVar(); self.tgt_path=tk.StringVar()
        self.src_rows=[]; self.tgt_rows=[]; self.matches=[]; self.src_maps=None
        self.sce=tk.IntVar(value=0); self.sca=tk.IntVar(value=1)
        self.scr=tk.IntVar(value=3); self.scrrc=tk.IntVar(value=5); self.sskip=tk.IntVar(value=1)
        self.tce=tk.IntVar(value=1); self.tca=tk.IntVar(value=2)
        self.tcn=tk.IntVar(value=3); self.tcv=tk.IntVar(value=4)
        self.tcrrc=tk.IntVar(value=5); self.tca2=tk.IntVar(value=10); self.tskip=tk.IntVar(value=1)
        self._man_results=[]; self._build()

    def _build(self):
        s=ttk.Style(self); s.theme_use("clam")
        s.configure("Treeview",background=BG3,foreground=TEXT,fieldbackground=BG3,rowheight=26,font=FN,borderwidth=0)
        s.configure("Treeview.Heading",background=BG2,foreground=ACCENT,font=FB,relief="flat")
        s.map("Treeview",background=[("selected",ACCENT2)])
        s.configure("T.Horizontal.TProgressbar",troughcolor=BG3,background=ACCENT,thickness=6)

        top=tk.Frame(self,bg=BG2,pady=8); top.pack(fill="x")
        tk.Label(top,text="⚡ Excel Data Matcher",font=FT,bg=BG2,fg=ACCENT).pack(side="left",padx=16)

        fp=tk.Frame(top,bg=BG2); fp.pack(side="left",padx=10)
        for r,(lbl,var,cmd) in enumerate([("Источник:",self.src_path,self._pick_src),
                                           ("Цель:",self.tgt_path,self._pick_tgt)]):
            tk.Label(fp,text=lbl,font=FB,bg=BG2,fg=TEXT_MUTE,width=10,anchor="e").grid(row=r,column=0,padx=4,pady=3,sticky="e")
            tk.Entry(fp,textvariable=var,width=55,font=FM,bg=BG3,fg=TEXT,insertbackground=ACCENT,relief="flat",bd=4
                     ).grid(row=r,column=1,padx=2,pady=3)
            tk.Button(fp,text="…",command=cmd,font=FB,bg=ACCENT2,fg="white",relief="flat",padx=8,cursor="hand2"
                      ).grid(row=r,column=2,padx=4,pady=3)

        bf=tk.Frame(top,bg=BG2); bf.pack(side="right",padx=14)
        for txt,cmd,col in [("⚙ Настройки",self._settings,ACCENT2),
                              ("▶ Автосопоставление",self._run,ACCENT),
                              ("💾 Сохранить",self._save,GREEN)]:
            tk.Button(bf,text=txt,command=cmd,font=FB,bg=col,fg="white",relief="flat",
                      padx=10,pady=4,cursor="hand2").pack(pady=2,fill="x")

        self.sv=tk.StringVar(value="Выберите файлы и нажмите Автосопоставление")
        tk.Label(self,textvariable=self.sv,font=FS,bg=BG,fg=TEXT_MUTE,anchor="w").pack(fill="x",padx=12,pady=(4,0))
        self.pb=ttk.Progressbar(self,mode="determinate"); self.pb.pack(fill="x",padx=12,pady=(2,4))

        pw=tk.PanedWindow(self,orient="horizontal",bg=BG,sashwidth=6); pw.pack(fill="both",expand=True,padx=6,pady=(0,6))
        lf=tk.Frame(pw,bg=BG); pw.add(lf,minsize=750)
        rf=tk.Frame(pw,bg=BG2,width=340); pw.add(rf,minsize=300)
        self._build_table(lf); self._build_manual(rf)

    def _build_table(self,p):
        tk.Label(p,text="Результаты сопоставления",font=FT,bg=BG,fg=TEXT).pack(anchor="w",padx=8,pady=(6,2))
        fb=tk.Frame(p,bg=BG); fb.pack(fill="x",padx=8,pady=(0,4))
        self.fc=ttk.Combobox(fb,values=["Все","Совпало","Не совпало","Нечёткое"],state="readonly",width=16,font=FS)
        self.fc.set("Все"); self.fc.pack(side="left",padx=4)
        self.fc.bind("<<ComboboxSelected>>",lambda _:self._filter())
        tk.Label(fb,text="Поиск:",font=FS,bg=BG,fg=TEXT_MUTE).pack(side="left",padx=(8,2))
        self.sq=tk.StringVar(); self.sq.trace_add("write",lambda *_:self._filter())
        tk.Entry(fb,textvariable=self.sq,width=28,font=FN,bg=BG3,fg=TEXT,insertbackground=ACCENT,relief="flat",bd=3).pack(side="left")
        self.statsv=tk.StringVar(); tk.Label(fb,textvariable=self.statsv,font=FS,bg=BG,fg=TEXT_MUTE).pack(side="right",padx=8)

        cols=("st","row","name","vol","ean","art","rrc","reason")
        self.tv=ttk.Treeview(p,columns=cols,show="headings",selectmode="browse")
        hdrs={"st":("",35),"row":("№",45),"name":("Название",280),"vol":("Объём",90),
              "ean":("EAN",135),"art":("Артикул",140),"rrc":("РРЦ",70),"reason":("Метод",170)}
        for c,(h,w) in hdrs.items():
            self.tv.heading(c,text=h); self.tv.column(c,width=w,minwidth=25)
        vs=ttk.Scrollbar(p,orient="vertical",command=self.tv.yview)
        hs=ttk.Scrollbar(p,orient="horizontal",command=self.tv.xview)
        self.tv.configure(yscroll=vs.set,xscroll=hs.set)
        vs.pack(side="right",fill="y"); hs.pack(side="bottom",fill="x")
        self.tv.pack(fill="both",expand=True,padx=(8,0))
        self.tv.tag_configure("ok",background="#1a3028",foreground=GREEN)
        self.tv.tag_configure("fail",background="#3a1a1a",foreground=RED)
        self.tv.tag_configure("fuzzy",background="#3a2a1a",foreground=ORANGE)
        self.tv.tag_configure("manual",background="#1a253a",foreground=ACCENT)
        self.tv.bind("<<TreeviewSelect>>",self._on_sel)

    def _build_manual(self,p):
        tk.Label(p,text="Ручное сопоставление",font=FT,bg=BG2,fg=TEXT).pack(anchor="w",padx=10,pady=(10,2))
        tk.Label(p,text="Выбрано:",font=FS,bg=BG2,fg=TEXT_MUTE).pack(anchor="w",padx=10)
        self.sl=tk.Label(p,text="—",font=FB,bg=BG2,fg=TEXT,wraplength=290,justify="left"); self.sl.pack(anchor="w",padx=10,pady=(0,8))
        tk.Label(p,text="Поиск в источнике:",font=FS,bg=BG2,fg=TEXT_MUTE).pack(anchor="w",padx=10)
        self.ms=tk.StringVar(); self.ms.trace_add("write",lambda *_:self._msearch())
        tk.Entry(p,textvariable=self.ms,width=34,font=FN,bg=BG3,fg=TEXT,insertbackground=ACCENT,relief="flat",bd=4).pack(fill="x",padx=10,pady=(2,6))
        self.ml=tk.Listbox(p,bg=BG3,fg=TEXT,font=FS,selectbackground=ACCENT2,relief="flat",activestyle="none",height=14)
        sb=ttk.Scrollbar(p,orient="vertical",command=self.ml.yview)
        self.ml.configure(yscrollcommand=sb.set)
        sb.pack(side="right",fill="y",padx=(0,10)); self.ml.pack(fill="both",expand=True,padx=10,pady=2)
        self.ml.bind("<Double-Button-1>",self._apply_man); self.ml.bind("<Return>",self._apply_man)
        tk.Label(p,text="Двойной клик / Enter — применить",font=FS,bg=BG2,fg=TEXT_MUTE).pack(pady=2)
        bf=tk.Frame(p,bg=BG2); bf.pack(fill="x",padx=10,pady=6)
        tk.Button(bf,text="Применить выбор",command=self._apply_man,font=FB,bg=ACCENT,fg="white",relief="flat",pady=4,cursor="hand2").pack(fill="x",pady=2)
        tk.Button(bf,text="Сбросить строку",command=self._clear,font=FB,bg=RED,fg="white",relief="flat",pady=4,cursor="hand2").pack(fill="x",pady=2)

    def _pick_src(self):
        p=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xls")])
        if p: self.src_path.set(p)
    def _pick_tgt(self):
        p=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xls")])
        if p: self.tgt_path.set(p)

    def _run(self):
        if not self.src_path.get() or not self.tgt_path.get():
            messagebox.showwarning("Файлы не выбраны","Выберите оба файла."); return
        self.pb["value"]=0; self.sv.set("Загружаю файлы..."); self.update_idletasks()
        threading.Thread(target=self._do_match,daemon=True).start()

    def _do_match(self):
        try:
            rows,am,em,nm,nom,wm=load_source(self.src_path.get(),self.sce.get(),self.sca.get(),self.scr.get(),self.scrrc.get(),self.sskip.get())
            self.src_rows=rows; self.src_maps=(am,em,nm,nom,wm)
            self.tgt_rows=load_target(self.tgt_path.get(),self.tce.get(),self.tca.get(),self.tcn.get(),self.tcv.get(),self.tcrrc.get(),self.tca2.get(),self.tskip.get())
            total=len(self.tgt_rows); self.matches=[]
            for i,tr in enumerate(self.tgt_rows):
                ent,reason=auto_match(tr,am,em,nm,nom,wm)
                self.matches.append([tr,ent,reason])
                self.after(0,lambda v=(i+1)/total*100:self._setp(v))
            self.after(0,self._pop_table)
        except Exception as e:
            self.after(0,lambda:messagebox.showerror("Ошибка",str(e)))

    def _setp(self,v): self.pb["value"]=v; self.update_idletasks()

    def _pop_table(self):
        self.tv.delete(*self.tv.get_children())
        ok=fail=fz=0
        for i,(tr,ent,reason) in enumerate(self.matches):
            st,tag=self._tag(ent,reason)
            if tag=="ok": ok+=1
            elif tag=="fail": fail+=1
            else: fz+=1
            self.tv.insert("","end",iid=str(i),tags=(tag,),values=(
                st,tr["row_idx"],tr["name"][:60],str(tr["vol"] or ""),
                ent["ean"] if ent else "","",ent["rrc"] if ent else "",reason if ent else "—"
            ))
            if ent:
                self.tv.set(str(i),"art",ent.get("article","") or "")
        total=len(self.matches)
        self.statsv.set(f"Совпало:{ok}  Нечёткое:{fz}  Нет:{fail}  /  {total}")
        self.sv.set(f"Готово! Совпало: {ok+fz}/{total}, не найдено: {fail}")
        self.pb["value"]=100

    def _tag(self,ent,reason):
        if not ent: return "❌","fail"
        if any(w in reason.lower() for w in ["fuzzy","keyword","нечёт","ключ"]): return "⚠","fuzzy"
        return "✅","ok"

    def _filter(self):
        if not self.matches: return
        fc=self.fc.get(); sq=self.sq.get().lower()
        for i,(tr,ent,reason) in enumerate(self.matches):
            _,tag=self._tag(ent,reason)
            show=True
            if fc=="Совпало" and tag!="ok": show=False
            if fc=="Не совпало" and tag!="fail": show=False
            if fc=="Нечёткое" and tag!="fuzzy": show=False
            if sq and sq not in tr["name"].lower():
                if not (ent and (sq in str(ent.get("article","") or "").lower() or sq in str(ent.get("ean","") or "").lower())): show=False
            item=str(i)
            if show:
                if not self.tv.exists(item): self.tv.reattach(item,"","end")
            else:
                if self.tv.exists(item): self.tv.detach(item)

    def _on_sel(self,_=None):
        sel=self.tv.selection()
        if not sel: return
        i=int(sel[0]); tr,ent,reason=self.matches[i]
        name_val = tr["name"]; vol_val = str(tr.get("vol","") or "")
        self.sl.config(text=name_val+"\n"+vol_val)
        self.ms.set(tr["name"]); self._msearch()

    def _msearch(self):
        q=self.ms.get().lower()
        kw=kwords(q); res=[]
        for row in self.src_rows:
            ru=(row.get("ru_name") or "").lower()
            sk=kwords(ru)
            sc=jaccard(kw,sk) if kw and sk else (0.1 if q in ru else 0)
            if sc>0.15 or (q and q[:6] in ru): res.append((sc,row))
        res.sort(key=lambda x:-x[0])
        self.ml.delete(0,"end"); self._man_results=[]
        for _,row in res[:80]:
            self.ml.insert("end",f"{row.get('ru_name','')}  |  {row.get('article','')}  |  РРЦ:{row.get('rrc','')}")
            self._man_results.append(row)

    def _apply_man(self,_=None):
        st=self.tv.selection(); sl=self.ml.curselection()
        if not st: messagebox.showinfo("Подсказка","Выберите строку в таблице."); return
        if not sl: messagebox.showinfo("Подсказка","Выберите запись в списке."); return
        i=int(st[0]); src=self._man_results[sl[0]]; tr=self.matches[i][0]
        self.matches[i]=[tr,src,"Вручную"]
        self.tv.item(str(i),values=("✏",tr["row_idx"],tr["name"][:60],str(tr["vol"] or ""),
                                     src["ean"],src["article"],src["rrc"],"Вручную"),tags=("manual",))

    def _clear(self):
        sel=self.tv.selection()
        if not sel: return
        i=int(sel[0]); tr=self.matches[i][0]
        self.matches[i]=[tr,None,""]
        self.tv.item(str(i),values=("❌",tr["row_idx"],tr["name"][:60],str(tr["vol"] or ""),"","","","—"),tags=("fail",))

    def _save(self):
        if not self.matches: messagebox.showinfo("Нет данных","Сначала выполните сопоставление."); return
        out=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")],
            initialfile=os.path.splitext(os.path.basename(self.tgt_path.get()))[0]+"_matched.xlsx",
            initialdir=os.path.dirname(self.tgt_path.get()))
        if not out: return
        try:
            wb=openpyxl.load_workbook(self.tgt_path.get()); ws=wb.worksheets[0]; filled=0
            for tr,ent,reason in self.matches:
                if not ent: continue
                ri=tr["row_idx"]
                bc=self.tce.get()+1; cc=self.tca.get()+1; fc=self.tcrrc.get()+1
                if not ws.cell(ri,bc).value and ent.get("ean"): ws.cell(ri,bc).value=ent["ean"]
                if not ws.cell(ri,cc).value and ent.get("article"): ws.cell(ri,cc).value=ent["article"]
                if not ws.cell(ri,fc).value and ent.get("rrc") is not None: ws.cell(ri,fc).value=ent["rrc"]
                filled+=1
            wb.save(out); messagebox.showinfo("Сохранено",f"Заполнено {filled} строк.\nФайл: {out}")
        except Exception as e: messagebox.showerror("Ошибка",str(e))

    def _settings(self):
        dlg=tk.Toplevel(self); dlg.title("Настройки колонок"); dlg.configure(bg=BG)
        dlg.geometry("540x480"); dlg.resizable(False,False); dlg.grab_set()
        def sec(txt):
            tk.Label(dlg,text=txt,font=FB,bg=BG,fg=ACCENT).pack(anchor="w",padx=16,pady=(10,0))
            tk.Frame(dlg,bg=BG3,height=1).pack(fill="x",padx=16)
        def rs(lbl,var):
            f=tk.Frame(dlg,bg=BG); f.pack(fill="x",padx=24,pady=2)
            tk.Label(f,text=lbl,font=FN,bg=BG,fg=TEXT,width=30,anchor="w").pack(side="left")
            tk.Label(f,text="(0=A 1=B...)",font=FS,bg=BG,fg=TEXT_MUTE).pack(side="right",padx=4)
            tk.Spinbox(f,from_=0,to=30,textvariable=var,width=5,bg=BG3,fg=TEXT,relief="flat",font=FN).pack(side="right")
        sec("Источник — колонки (0-индекс)")
        for l,v in [("EAN",self.sce),("Артикул",self.sca),("Рус. название",self.scr),("РРЦ",self.scrrc),("Пропустить строк",self.sskip)]: rs(l,v)
        sec("Цель — колонки (0-индекс)")
        for l,v in [("EAN",self.tce),("Артикул",self.tca),("Название",self.tcn),("Объём",self.tcv),("РРЦ",self.tcrrc),("Артикул 2",self.tca2),("Пропустить строк",self.tskip)]: rs(l,v)
        tk.Button(dlg,text="Закрыть",command=dlg.destroy,font=FB,bg=ACCENT,fg="white",relief="flat",pady=6,cursor="hand2").pack(pady=12,padx=16,fill="x")

if __name__=="__main__":
    App().mainloop()
